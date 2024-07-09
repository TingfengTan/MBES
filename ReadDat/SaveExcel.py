import openpyxl
import numpy as np
import pynmea2

class SaveExcel():
    def __init__(self, Dat) -> None:
        self.workbook1 = openpyxl.Workbook()
        self.workbook2 = openpyxl.Workbook()
        self.workbook3 = openpyxl.Workbook()
        self.workbook4 = openpyxl.Workbook()
        self.Dat = Dat

    def save(self):
        # save AllDepth
        sheet_name = "AllDepth"
        if sheet_name not in self.workbook1.sheetnames:
            sheet1 = self.workbook1.create_sheet(sheet_name)
            keys_list = list(self.Dat.AllDepth[0].keys())
            for i in range(len(keys_list)-1):
                sheet1.cell(row=1,column=i+1).value = keys_list[i]
            sheet1_sample_number = self.workbook1.create_sheet("AllDepth_sample_number")
            sheet1_angle = self.workbook1.create_sheet("AllDepth_angle")
            sheet1_upper_gate = self.workbook1.create_sheet("AllDepth_upper_gate")
            sheet1_lower_gate = self.workbook1.create_sheet("AllDepth_lower_gate")
            sheet1_intensity = self.workbook1.create_sheet("AllDepth_intensity")
            sheet1_flags = self.workbook1.create_sheet("AllDepth_flags")
            sheet1_quality_flag = self.workbook1.create_sheet("AllDepth_quality_flag")
            sheet1_quality_val = self.workbook1.create_sheet("AllDepth_quality_val")
            for i in range(self.Dat.AllDepth[0]['N']):
                sheet1_sample_number.cell(row=1,column=i+1).value = i+1
                sheet1_angle.cell(row=1,column=i+1).value = i+1
                sheet1_upper_gate.cell(row=1,column=i+1).value = i+1
                sheet1_lower_gate.cell(row=1,column=i+1).value = i+1
                sheet1_intensity.cell(row=1,column=i+1).value = i+1
                sheet1_flags.cell(row=1,column=i+1).value = i+1
                sheet1_quality_flag.cell(row=1,column=i+1).value = i+1
                sheet1_quality_val.cell(row=1,column=i+1).value = i+1
        else:
            sheet1 = self.workbook1[sheet_name]
        for row, item in enumerate(self.Dat.AllDepth, start=2):
            for col, key in enumerate(item.keys(), start=1):
                if key == 'data':
                    pass
                    N = item[key].shape[0]
                    for j in range(N):
                        sheet1_sample_number.cell(row = row,column = j+1).value = item[key][j,0]
                        sheet1_angle.cell(row = row,column = j+1).value = item[key][j,1]
                        sheet1_upper_gate.cell(row = row,column = j+1).value = item[key][j,2]
                        sheet1_lower_gate.cell(row = row,column = j+1).value = item[key][j,3]
                        sheet1_intensity.cell(row = row,column = j+1).value = item[key][j,4]
                        sheet1_flags.cell(row = row,column = j+1).value = item[key][j,5]
                        sheet1_quality_flag.cell(row = row,column = j+1).value = item[key][j,6]
                        sheet1_quality_val.cell(row = row,column = j+1).value = item[key][j,7]
                else:
                    sheet1.cell(row=row, column=col).value = item[key]
        self.workbook1.save("./data/AllDepth.xlsx")


        # save AllWater
        sheet_name = "AllWater"
        if sheet_name not in self.workbook2.sheetnames:
            sheet2 = self.workbook2.create_sheet(sheet_name)
            keys_list = list(self.Dat.AllWater[0].keys())
            for i in range(len(keys_list)-2):
                sheet2.cell(row=1,column=i+1).value = keys_list[i]
            sheet2_bmangle = self.workbook2.create_sheet("AllWater_bmangle")
            for i in range(self.Dat.AllWater[1]['N']):
                sheet2_bmangle.cell(row=1,column=i+1).value = i+1
        else:
            sheet2 = self.workbook[sheet_name]
        npy = []
        for row, item in enumerate(self.Dat.AllWater, start=2):
            for col, key in enumerate(item.keys(), start=1):
                if key == 'data':
                    N = item[key].shape[0]
                    npy.append(item[key])
                    
                elif key=='bmAng':
                    N = item[key].shape[0]
                    for j in range(N):
                        sheet2_bmangle.cell(row=row,column=j+1).value = item[key][j]
                    
                else:
                    sheet2.cell(row=row, column=col).value = item[key] 
        self.workbook2.save("./data/AllWater.xlsx")
        np.savez('./data/AllWater_Data.npz',*npy)

        # save MRU
        sheet_name = 'MRU'
        if sheet_name not in self.workbook3.sheetnames:
            sheet3 = self.workbook3.create_sheet(sheet_name)
            sheet3.cell(row=1,column=1).value = "Horizontal Acceleration"
            sheet3.cell(row=1,column=2).value = "Vertical Acceleration"
            sheet3.cell(row=1,column=3).value = "Heave Polarity"
            sheet3.cell(row=1,column=4).value = "Heave"
            sheet3.cell(row=1,column=5).value = "Status Flag"
            sheet3.cell(row=1,column=6).value = "Roll Polarity"
            sheet3.cell(row=1,column=7).value = "Roll"
            sheet3.cell(row=1,column=8).value = "Pitch Polarity"
            sheet3.cell(row=1,column=9).value = "Pitch"
        else:
            sheet3 = self.workbook[sheet_name]

        for row, item in enumerate(self.Dat.MRU, start=2):
            HA = int(item[1:3],16) * 0.0383
            VA = int(item[3:7],16) * 0.000625 - 20.48
            HP = item[8]
            H = int(item[9:13],16) * 0.01
            SF = item[13]
            RP = item[14]
            R = int(item[15:19],16) * 0.01
            PP = item[20]
            P = int(item[21:25], 16) * 0.01
            sheet3.cell(row=row,column=1).value = HA
            sheet3.cell(row=row,column=2).value = VA
            sheet3.cell(row=row,column=3).value = HP
            sheet3.cell(row=row,column=4).value = H
            sheet3.cell(row=row,column=5).value = SF
            sheet3.cell(row=row,column=6).value = RP
            sheet3.cell(row=row,column=7).value = R
            sheet3.cell(row=row,column=8).value = PP
            sheet3.cell(row=row,column=9).value = P
        self.workbook3.save("./data/MRU.xlsx")

        #  save GPS
        sheet_name = 'GPS'
        if sheet_name not in self.workbook4.sheetnames:
            sheet4 = self.workbook4.create_sheet(sheet_name)
            sheet4.cell(row=1,column=1).value = "heading"  # HDT
            sheet4.cell(row=1,column=2).value = "hdg_true" # HDT
            
            sheet4.cell(row=1,column=3).value = "true_track" #VTG
            sheet4.cell(row=1,column=4).value = "true_track_sym" #VTG
            sheet4.cell(row=1,column=5).value = "mag_track" #VTG
            sheet4.cell(row=1,column=6).value = "mag_track_sym" #VTG
            sheet4.cell(row=1,column=7).value = "spd_over_grnd_kts" #VTG
            sheet4.cell(row=1,column=8).value = "spd_over_grnd_kts_sym" #VTG
            sheet4.cell(row=1,column=9).value = "spd_over_grnd_kmph" #VTG
            sheet4.cell(row=1,column=10).value = "spd_over_grnd_kmph_sym" #VTG
            sheet4.cell(row=1,column=11).value = "faa_mode" #VTG
            
            sheet4.cell(row=1,column=12).value = "timestamp" #ZDA
            sheet4.cell(row=1,column=13).value = "day" #ZDA
            sheet4.cell(row=1,column=14).value = "month" #ZDA
            sheet4.cell(row=1,column=15).value = "year" #ZDA
            sheet4.cell(row=1,column=16).value = "local_zone" #ZDA
            sheet4.cell(row=1,column=17).value = "local_zone_minutes" #ZDA
            
            sheet4.cell(row=1,column=18).value = "timestamp" #GGA
            sheet4.cell(row=1,column=19).value = "lat" #GGA
            sheet4.cell(row=1,column=20).value = "lat_dir" #GGA
            sheet4.cell(row=1,column=21).value = "lon" #GGA
            sheet4.cell(row=1,column=22).value = "lon_dir" #GGA
            sheet4.cell(row=1,column=23).value = "gps_qual" #GGA
            sheet4.cell(row=1,column=24).value = "num_sats" #GGA
            sheet4.cell(row=1,column=25).value = "horizontal_dil" #GGA
            sheet4.cell(row=1,column=26).value = "altitude" #GGA
            sheet4.cell(row=1,column=27).value = "altitude_units" #GGA
            sheet4.cell(row=1,column=28).value = "geo_sep" #GGA
            sheet4.cell(row=1,column=29).value = "geo_sep_units" #GGA
            sheet4.cell(row=1,column=30).value = "age_gps_data" #GGA
            sheet4.cell(row=1,column=31).value = "ref_station_id" #GGA
        else:
            sheet4 = self.workbook[sheet_name]

        start = 1
        HDT_flag = 0
        VTG_flag = 2
        ZDA_flag = 11
        GGA_flag = 17
        for (_, item) in enumerate(self.Dat.GPS):
            flag = item[:6]
            if flag == '$GPGGA':
                tmp = pynmea2.parse(item[:-2])
                sheet4.cell(row=start, column=GGA_flag+1).value = tmp.timestamp.strftime("%H:%M:%S")
                sheet4.cell(row=start, column=GGA_flag+2).value = tmp.lat
                sheet4.cell(row=start, column=GGA_flag+3).value = tmp.lat_dir
                sheet4.cell(row=start, column=GGA_flag+4).value = tmp.lon
                sheet4.cell(row=start, column=GGA_flag+5).value = tmp.lon_dir
                sheet4.cell(row=start, column=GGA_flag+6).value = tmp.gps_qual
                sheet4.cell(row=start, column=GGA_flag+7).value = tmp.num_sats
                sheet4.cell(row=start, column=GGA_flag+8).value = tmp.horizontal_dil
                sheet4.cell(row=start, column=GGA_flag+9).value = tmp.altitude
                sheet4.cell(row=start, column=GGA_flag+10).value = tmp.altitude_units
                sheet4.cell(row=start, column=GGA_flag+11).value = tmp.geo_sep
                sheet4.cell(row=start, column=GGA_flag+12).value = tmp.geo_sep_units
                sheet4.cell(row=start, column=GGA_flag+13).value = tmp.age_gps_data
                sheet4.cell(row=start, column=GGA_flag+14).value = tmp.ref_station_id
            elif flag == '$GPVTG':
                tmp = pynmea2.parse(item[:-2])
                sheet4.cell(row=start, column=VTG_flag+1).value = tmp.true_track
                sheet4.cell(row=start, column=VTG_flag+2).value = tmp.true_track_sym
                sheet4.cell(row=start, column=VTG_flag+3).value = tmp.mag_track
                sheet4.cell(row=start, column=VTG_flag+4).value = tmp.mag_track_sym
                sheet4.cell(row=start, column=VTG_flag+5).value = tmp.spd_over_grnd_kts
                sheet4.cell(row=start, column=VTG_flag+6).value = tmp.spd_over_grnd_kts_sym
                sheet4.cell(row=start, column=VTG_flag+7).value = tmp.spd_over_grnd_kmph
                sheet4.cell(row=start, column=VTG_flag+8).value = tmp.spd_over_grnd_kmph_sym
                sheet4.cell(row=start, column=VTG_flag+9).value = tmp.faa_mode
            elif flag == '$GPZDA':
                tmp = pynmea2.parse(item[:-2])
                sheet4.cell(row=start, column=ZDA_flag+1).value = tmp.timestamp.strftime("%H:%M:%S")
                sheet4.cell(row=start, column=ZDA_flag+2).value = tmp.day
                sheet4.cell(row=start, column=ZDA_flag+3).value = tmp.month
                sheet4.cell(row=start, column=ZDA_flag+4).value = tmp.year
                sheet4.cell(row=start, column=ZDA_flag+5).value = tmp.local_zone
                sheet4.cell(row=start, column=ZDA_flag+6).value = tmp.local_zone_minutes
            elif flag == '$GPHDT': # GPHDT
                start = start+1
                tmp = pynmea2.parse(item[:-2])
                sheet4.cell(row=start, column=HDT_flag+1).value = tmp.heading
                sheet4.cell(row=start, column=HDT_flag+2).value = tmp.hdg_true
                
        self.workbook4.save("./data/GPS.xlsx")



        


        